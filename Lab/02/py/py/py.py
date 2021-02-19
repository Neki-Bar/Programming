
import openpyxl
import datetime
import json
import os.path
from flask import Flask, request

app = Flask(__name__)

def write_from_buf():
    global buf, next_line
    book = openpyxl.load_workbook('data.xlsx')
    sheet = book.active
    
    for line in buf:
        for i in range(5):
            sheet.cell(next_line, i + 1).value = line[i]
        next_line += 1

    sheet.cell(2, 6).value = next_line - 1
    book.save('data.xlsx')
    book.close
    buf.clear()

@app.route('/', methods = ['POST'])
def index():
    global id, buf
    time = datetime.datetime.now().time()

    req = request.get_json()
    
    for item in req:
        next_line = [
            id,
            req["user_id"],
            time,
            item["item"],
            item["price"]
        ]
        id += 1
        buf.append(line)
       

    if len(buf) > 1:
        write_from_buf()

if __name__ == "__main__":
    global buf, id, next_line
    id   = 1
    next_line = 2
    buf       = []

    if not os.path.exists('data.xlsx'):
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.cell(1,1).value = 'N'
        sheet.cell(1,2).value = 'User ID'
        sheet.cell(1,3).value = 'Datetime'
        sheet.cell(1,4).value = 'Item'
        sheet.cell(1,5).value = 'Price'
        sheet.cell(1,6).value = 'Amount of Items'
        sheet.cell(2,6).value = 1
        book.save('data.xlsx')
        book.close
    else:
        book = openpyxl.load_workbook('data.xlsx')
        sheet = book.active

        id = int(sheet.cell(2,6).value)
        next_line = id + 1

        book.close

    app.run()